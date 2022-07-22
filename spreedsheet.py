import openpyxl
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file("sheet1")
"""list each company woith respective products count
   output:{'AAA COMPANY:43, 'BBB COMPANY:17', 'CCC COMPANY:14'} """
product_per_supplier = {}
for product_row in range(2, product_list.max_row):
    supplier_name = product_list.cell(product_row, 4).value
    current_num_products = product_per_supplier.get(supplier_name)
    product_per_supplier[supplier_name] = current_num_products+1
else:
    product_per_supplier[supplier_name] = 1
print(product_per_supplier)

